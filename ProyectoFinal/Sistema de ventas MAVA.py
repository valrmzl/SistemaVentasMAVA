import openpyxl
from datetime import date
from colorama import init, Fore,Style,Back
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('base.xlsx')

def sesion(p1, p2):   #función que me permite hacer la validación de usuario y contraseña
    wb.active = 1
    ws = wb.active  # variable
    usuario = ws["A"]
    contraseña = ws["B"]
    u = []
    c = []
    j=0
    for U in usuario:  # ciclo que me añade a una lista todos los usuarios
        u.append(U.value)
    for C in contraseña:  # ciclo que me añade a una lista todas las contraseñas
        c.append(C.value)
    u.pop(0)
    c.pop(0)
    if usu in u:  #busco que el usuario se encuentre en la lista correspondiente
        for i in u: #de ser asi este contador ayuda a saber la posiicón que tiene en u (donde i corresponde al usuario)
            j=j+1
            if usu==i:    #si son iguales continua
                break
        if contra==c[j-1]: #como la lista empieza en 0 se le resta uno para saber si sus posiciones coinciden
            return 1       #de ser correcto regresa 1 para que se haga valida la funcion de iniciar sesion
        else:
            return 2 
    else:
        return 2

def menup():  # función menu principal
    init(autoreset=True)
    print(Fore.MAGENTA + Style.BRIGHT + "*****************************\n**** B I E N V E N I D O ****\n** Sistema de  Ventas MAVA **\n*****************************")
    print(
        Style.BRIGHT + "Opciones disponibles: \n1.-Control de almacén.\n2.-Directorio de clientes.\n3.-Ventas.\n4.-Estadísticas.\n5.-Configuración de usuarios.\n6.-Salir.")
    while True:
        try:
            op = int(input(Fore.CYAN + Style.BRIGHT + "Seleccione una opción: "))
            if op in range(1, 7):
                break
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 6.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números.")
    return op

def mca():  # función menu control de almacen
    init(autoreset=True)
    while True:
        print(Style.BRIGHT + "Bienvenido al Control de Almacén\n1.-Alta de producto\n2.-Eliminar producto\n3.-Consulta de producto\n4.-Añadir producto almacen\n5.-Salir")
        try:
            op = int(input(Fore.CYAN + Style.BRIGHT + "Seleccione una opción: "))
            if op in range(1, 6):
                return op
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 5.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números")

def anadir_producto():  # función añadir productos al almacen
    info = list([])
    wb.active = 0
    sheet = wb.active
    print(Fore.YELLOW + "---Alta de producto---")
    print("Introduzca los datos correspondientes.")
    print('No. de producto: ', sheet.max_row)
    while True:
        try:
            can =float(input("Cantidad: "))
            if can>0:
                break
            else:
                print(Fore.RED+"Introduce una cantidad válida")
        except ValueError:
            print(Fore.RED+"Error deben de ser cantidades")
    info.append(can)
    pre = input("Presentación: ")
    info.append(pre.upper())
    nombre = input("Nombre del producto: ")
    info.append(nombre.upper())
    while True:
        try:
            pc = float(input("Precio de compra: $"))
            if pc>0:
                break
            else:
                print(Fore.RED+"Introduce una cantidad válida")
        except ValueError:
            print(Fore.RED+"Error deben de ser cantidades")
    info.append(pc)
    while True:
        try:
            pv = float(input("Precio de venta: $"))
            if pv>0:
                break
            else:
                print(Fore.RED+"Introduce una cantidad válida")
        except ValueError:
            print(Fore.RED+"Error deben de ser cantidades")
    info.append(pv)
    print(Fore.YELLOW + "Producto añadido con éxito\n*******")
    sheet.append(info)  # añade lista con información del producto a excel
    wb.save("base.xlsx")
    wb.close()


def eliminar_producto():  # función para eliminar productos del alamcen
    print(Fore.YELLOW + '---Eliminar producto---')
    wb = openpyxl.load_workbook('base.xlsx')
    busc = list([])
    wb.active = 0
    ws = wb.active
    productos = ws["C"]
    for i in productos:
        busc.append(i.value)  # lista con nombres de productos
    while True:
        try:
            p = input('Producto: ')
            p = p.upper()
            no = 0
            if p in busc:  # busca input en lista de productos
                no = busc.index(p) + 1
                print('El No. de producto es: ', no - 1, )
                print("¿Seguro que quieres eliminar el producto", p, "con número", no - 1, "? si/no ", end="")
                ans = input()
                ans = ans.upper()
                if ans == "SI":
                    ws.delete_rows(no)  # eliminar producto
                    print(Fore.YELLOW + "Producto eliminado con éxito")
                    break
                else:
                    break
            else:
                print('Producto no encontrado. Intente de nuevo.')
        except ValueError:
            print('Producto no encontrado. Intente de nuevo.')
    wb.save("base.xlsx")
    wb.close()

def compra_producto():  # funcion para aumentar la cantidad de un producto en existencia en el almacen
    print(Fore.YELLOW + "---Compra producto, añadir almacen---\n¿Que producto va a añadir en almacen?")
    wb = openpyxl.load_workbook('base.xlsx')
    busc = list([])
    wb.active = 0
    ws = wb.active
    productos = ws["C"]
    for i in productos:
        busc.append(i.value)  # añade a una lista todos los productos
    while True:
        try:
            p = input('Producto: ')
            p = p.upper()
            if p in busc:  # checa si el producto se encuentra en la lista
                no = busc.index(p) + 1
                np = float(input('Cantidad: '))
                ws.cell(row=no, column=1).value = float(ws.cell(row=no, column=1).value) + np  # añade la cantidad de producto
                print('Producto añadido con éxito.')
                break
            else:
                print(Fore.RED + 'Producto no encontrado. Intente de nuevo.')
        except ValueError:
            print('Inserte un valor válido.')
    wb.save("base.xlsx")
    wb.close()


def menu_conprod():  # funcion de menu de consultar productos
    init(autoreset=True)
    while True:
        print("¿Que desea hacer?\n1.-Consultar productos\n2.-Salir")
        try:
            op = int(input("Seleccione una opción: "))
            if op in range(1, 3):
                return op
            else:
                print(Fore.RED + 'Error, ingrese números del 1 al 2.')
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números.")


def consultar_productos():  # funcion de consultar producto en almacen
    print(Fore.YELLOW + "---Consulta de Producto---")
    wb = openpyxl.load_workbook('base.xlsx')
    busc = list([])
    wb.active = 0
    ws = wb.active
    productos = ws["C"]
    for i in productos:
        busc.append(i.value)  # lista con todos los productos
    p = input('Producto: ')
    p = p.upper()
    if p in busc:
        no = busc.index(p) + 1
        prod = []
        for i in range(1, 7):
            prod.append(ws.cell(row=no, column=i).value)  # añade los datos del producto a una lista
        print('No. de producto:', no - 1, '\nExistencia:', prod[0], '\nPresentación:', prod[1], '\nProducto:', prod[2],
              '\nPrecio de venta: $', prod[4])
    else:
        print(Fore.RED + 'Producto no encontrado. Intente de nuevo.')
    wb.save("base.xlsx")
    wb.close()

def mdc():  # funcion menu del directorio
    init(autoreset=True)
    while True:
        print(Style.BRIGHT + "Bienvenido al Directorio de Clientes\n1.-Añadir cliente\n2.-Eliminar cliente\n3.-Salir")
        try:
            op = int(input(Style.BRIGHT + Fore.CYAN + "Seleccione una opción: "))
            if op in range(1, 4):
                return op
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 3.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números")


def anadir_contacto():  # función para añadir contactos
    init(autoreset=True)
    info = list([])
    wb.active = 2
    sheet = wb.active
    print(Fore.YELLOW + "---Añadir Contacto---")
    print("Introduzca los datos correspondientes.")
    print('No. de cliente', sheet.max_row)
    nombre = input("Nombre completo: ")
    info.append(nombre)
    tel = int(input("Teléfono: "))
    info.append(tel)
    correo = input("Correo electrónico: ")
    info.append(correo)
    rfc = input("RFC: ")
    info.append(rfc)
    dirf = input("Dirección Fiscal: ")
    info.append(dirf)
    dirop = input("Dirección Operativa: ")
    info.append(dirop)
    ncompras = float(0)
    mcompras = float(0)
    info.append(ncompras)
    info.append(mcompras)
    print(Fore.YELLOW + "Contacto añadido con éxito\n*******")
    sheet.append(info)  # añade datos del cliente a hoja de excel
    wb.save("base.xlsx")
    wb.close()
    return info  # regresa los datos del cliente

def eliminar_contacto():  # funcion para eliminar contactos
    init(autoreset=True)
    print(Fore.YELLOW + "---Eliminar contacto---")
    print("Esta es su lista de contactos: ")
    wb.active = 2
    ws = wb.active  # variable n= ws["A"]
    print("No.        Nombre cliente: ")  # no=[]
    a = 0
    for i in range(2, ws.max_row + 1):  # imprime la lista completa de clientes
        print(a + 1, end=" ")
        a = a + 1
        print("     ", ws.cell(row=i, column=1).value)
    while True:
        try:
            print("ESTE NUMERO ES SOLO DE REFERENICA PARA SU BORRADO, NO ES EL DE NO. DE CLIENTE REAL")
            borrar = int(input("No. de cliente que deseas borrar: "))
            if borrar in range(1, ws.max_row):
                ws.delete_rows(borrar + 1)  # elimina datos en excel
                print(Fore.YELLOW + "Borrado exitosamente")
                break
            else:
                print(Fore.RED + "Introduce un número válido.")
        except ValueError:
            print(Fore.RED + "Ingrese solo números.")
    wb.save("base.xlsx")
    wb.close()


def mventas():  # menu de ventas
    init(autoreset=True)
    while True:
        print(Style.BRIGHT + "Bienvenido a Ventas.\n1.-Nueva nota\n2.-Salir")
        try:
            op = int(input(Style.BRIGHT + Fore.CYAN + "Seleccione una opción: "))
            if op in range(1, 3):
                return op
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 3.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números.")


def nota_venta():  # funcon que realiza la nota de venta
    print(Fore.YELLOW + "***NOTA DE VENTA***")
    print("Fecha:", f)
    wb.active = 0
    ws = wb.active
    productos = ws["C"]
    busc = []
    for i in productos:
        busc.append(i.value)  # crea lista de productos para busqueda
    n = int(input('Número de productos en nota de ventas: '))
    nota = []
    prod = []
    i = 1  # contador para el ciclo
    total = 0
    while i <= n:  # ciclo para añadir los productos requeridos a la nota de ventas
        p = input('Producto: ')
        p = p.upper()
        if p in busc:
            no = busc.index(p)  # comprobar que el producto se encuentre en el inventario
            while True:
                try:
                    c = float(input('Cantidad: '))
                    i = i + 1
                    break
                except ValueError:
                    print('Introduce un número: ')
            precio = c * float(ws.cell(row=no + 1, column=5).value)
            precioc = c * float(ws.cell(row=no + 1, column=4).value)
            prod.append(c)
            prod.append(p)
            prod.append(precio)
            prod.append(precioc)
            nota.append(prod)  # se añaden los datos del producto a la nota
            prod = []
            total = total + float(precio)  # calcula el total
        else:
            print(Fore.RED + 'Producto no se encuentra en la base.')
    wb.save("base.xlsx")
    wb.close()
    return nota, total  # regresa nota con lista de productos y total


def seleccionar_cliente(total):  # funcion que me permite seleccionar cliente para ser usado en la nota de venta
    print(Fore.YELLOW + "Esta es su lista de contactos: ")
    wb.active = 2
    ws = wb.active  # variable n= ws["A"]
    print("No.        Nombre cliente: ")  # no=[]
    a = 1
    for i in range(1, ws.max_row):  # imprime lista de todos los clientes para seleccion
        print(a, end=" ")
        a = a + 1
        print("     ", ws.cell(row=i + 1, column=1).value)
    while True:
        try:  # comprueba que el cliente este en la lista
            nclient = int(input('Número del cliente: ')) + 1
            if nclient in range(1, ws.max_row + 1):
                client = str(ws.cell(row=nclient, column=1).value)
                break
            else:
                print(Fore.RED + "Introduce un número válido")
        except ValueError:
            print(Fore.RED + 'Ingrese solo números.')
    ws.cell(row=nclient, column=7).value = int(ws.cell(row=nclient, column=7).value) + 1  # añade otra compra a la celda del cliente
    ws.cell(row=nclient, column=8).value = float(ws.cell(row=nclient, column=8).value) + total  # añade el monto del gasto del cliente
    wb.save("base.xlsx")
    wb.close()
    return client  # regresa nombre del cliente


def imprimir_nota(n, c, t):  # función que me permite imprimir la nota de venta
    init(autoreset=True)
    wb.active = 3
    ws = wb.active
    productos = ws["B"]
    busc = []
    for i in productos:
        busc.append(i.value)  # añade productos de hoja nota de venta a una lista
    print(Fore.YELLOW + '****NOTA DE VENTA****')
    print(f)
    print('Cliente: ', c)
    for j in n:  # funcion que lee cada producto en la nota
        p = j
        print(p[0], p[1], p[2])  # imprime productos para cuenta
        if p[1] in busc:  # busca si el producto ya se vendio antes
            no = busc.index(p[1])  # fila en que esta el producto
            ws.cell(row=no + 1, column=1).value = float(ws.cell(row=no + 1, column=1).value) + p[0]  # suma cantidad vendida
            ws.cell(row=no + 1, column=3).value = float(ws.cell(row=no + 1, column=3).value) + p[2]  # suma monto vendido (precio compra)
            ws.cell(row=no + 1, column=4).value = float(ws.cell(row=no + 1, column=4).value) + p[3]  # suma monto vendido
        else:
            ws.append(p)  # añade el producto a hoja nota de ventas
    print('TOTAL SIN IVA: $', round(t * 0.84, 2))
    print("IVA: $", round(t * 0.16, 2))
    print(Style.BRIGHT + 'TOTAL: $', t)
    wb.save("base.xlsx")
    wb.close()

def resta_almacen(n):  # funcion que realiza los cambios en el almacen dependiendo de la nota de venta
    wb.active = 0
    ws = wb.active
    productos = ws["C"]
    busc = []
    for i in productos:
        busc.append(i.value)  # añade todos los productos a una lista
    for j in n:  # ciclo que recorre los productos en la nota de ventas
        p = j
        if p[1] in busc:  # si el producto esta en la lista del almacen
            no = busc.index(p[1])
            ws.cell(row=no + 1, column=1).value = float(ws.cell(row=no + 1, column=1).value) - p[0]  # resta cantidad de producto a almacen
        else:
            print(Fore.RED + 'Producto no se encuentra en la base.')
    wb.save("base.xlsx")
    wb.close()
    print(Style.BRIGHT + 'Gracias por su compra.')

def mestadisticas():  # menu estadisticas
    init(autoreset=True)
    while True:
        print(Style.BRIGHT + "Bienvenido al menú de Estadísticas.\n1.-Producto más vendido\n2.-Ganancias\n3.-Producto con poca existencia.\n4.-Cliente frecuente\n5.-Cliente mayor aportación\n6.-Salir")
        try:
            op = int(input(Style.BRIGHT + Fore.CYAN + "Seleccione una opción: "))
            if op in range(1, 7):
                return op  # regresa opcion para siguiente menu
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 6.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números")

def pvendidod():  # muestra los productos más vendidos segun el dinero
    print(Fore.YELLOW + 'PRODUCTO MAS VENDIDO (SEGÚN DINERO)')
    wb.active = 3
    ws = wb.active
    prod = ws['B']
    list = []
    a = int(1)  # contador para ir renglon por renglon
    art = []
    for i in prod:
        c = ws.cell(row=a, column=3).value  # celda de precio
        art.append(c)
        art.append(i.value)  # nombre de producto
        list.append(art)  # añade a una lista los articulos, cada articulo es una lista tambien
        art = []  # elimina datos de articulo para añadir otro
        a = a + 1
    list.pop(0)  # elimina el renglon de cantidad, producto y precio
    productos = []
    precios = []
    list.sort(reverse=True)  # ordena lista de productos de mayor a menor (monto vendido)
    a = 0  # contador
    for a in range(5):  # añade a una lista los productos en este caso solo top 5
        i = list[a]  # separa cada producto
        productos.append(i[0])  # lista con productos
        precios.append(i[1])  # lista con precios de productos
    plt.figure(figsize=(10, 3))
    plt.barh(precios, productos)  # grafica con valores de ambas listas
    plt.suptitle('Productos más vendidos ($$$):')
    plt.show()
    wb.save("base.xlsx")
    wb.close()


def ganancias_diaria():  # muestra las ganancias diarias
    wb.active = 3
    ws = wb.active
    dv = ws["C"]
    dc = ws["D"]
    dinv = []
    dinc = []
    for i in dv:
        dinv.append(i.value)  # valores de monto del producto vendido con precio de venta
    dinv.pop(0)
    for i in dc:
        dinc.append(i.value)  # valores de monto del producto vendido con precio de compra
    dinc.pop(0)
    dinerov = sum(dinv)  # suma todos valores
    dineroc = sum(dinc)  # suma todos valores
    ganancia = dinerov - dineroc  # monto vendido - precio de compra
    wb.save("base.xlsx")
    wb.close()
    print(Fore.YELLOW + 'La ganancia de las notas fue de: $', ganancia)
    return ganancia


def guardar_ganancia(g):  # función que guarda las ganancias por dia
    f = date.today()
    wb.active = 4
    ws = wb.active
    m = f, g  # lista con fecha y ganancia
    ws.append(m)  # guarda en excel
    wb.save("base.xlsx")
    wb.close()


def pexistencia():  # muestra los produtocs con menos existencia
    print(Fore.YELLOW + 'PRODUCTO CON POCA EXISTENCIA')
    wb.active = 0
    ws = wb.active
    cantidad = ws["A"]
    busc = []
    for i in cantidad:
        busc.append(i.value)  # añade a una lista todas las cantidades de producto
    busc.pop(0)  # elimina fila de cantidad, producto, precio
    a = 0  # contador
    n = []  # lista de cantidades
    p = []  # lista de productos
    for j in busc:  # recorre todos los productos
        a = a + 1
        if int(j) <= 10:  # si hay menos de 10 en almacen de dicho producto
            no = ws.cell(row=a + 1, column=1).value  # copia valor de cantidad
            prod = ws.cell(row=a + 1, column=3).value  # copia nombre del producto
            n.append(no)  # añade las menores cantidades a una lista
            p.append(prod)  # añade los productos a esa lista
    plt.figure(figsize=(5, 3))
    plt.barh(p, n)  # grafica con ambas listas, cantidades <10 y productos
    plt.suptitle('Productos con poca existencia')
    plt.show()
    wb.save("base.xlsx")
    wb.close()


def cliente_frecuente():  # muestra a los 5 cliente más frecuentes en relación de numero de compras
    print(Fore.YELLOW + 'CLIENTES FRECUENTES (5)')
    wb.active = 2
    ws = wb.active
    nombre = ws['A']
    list = []
    a = int(1)
    client = []
    for i in nombre:  # recorre toda la lista de clientes
        c = ws.cell(row=a, column=7).value  # toma valor de número de compras
        client.append(c)
        client.append(i.value)
        list.append(client)  # lista con numero de compras y nombre de cliente
        client = []
        a = a + 1  # contador para pasar por cada fila
    list.pop(0)  # elimina la fila de nombre, cantidad telefono, etc.
    list.pop(0)  # elimina la fila de publico general
    clientes = []
    compras = []
    list.sort(reverse=True)  # ordena los valores de mayor a menor
    for j in range(5):  # ciclo añade los 5 clientes con mayor numero de compras
        i = list[j]
        clientes.append(i[0])  # añade nombre de clientes
        compras.append(i[1])  # añade el número de compras
    plt.figure(figsize=(15, 3))
    plt.barh(compras, clientes)  # grafica con valores
    plt.suptitle('Clientes con mayor número de compras: ')
    plt.show()
    wb.save("base.xlsx")
    wb.close()

def clientes_dinero():  # clientes que más han consumido
    print(Fore.YELLOW + 'CLIENTES MAYOR APORTACIÓN (5)')
    wb.active = 2
    ws = wb.active
    nombre = ws['A']
    list = []
    a = int(1)
    client = []
    for i in nombre:  # ciclo que recorre todos los clientes
        c = ws.cell(row=a, column=8).value  # copia valor de monto de compra
        client.append(c)
        client.append(i.value)
        list.append(client)  # lista de clientes con monto de compras
        client = []
        a = a + 1  # contador para recorrer cada fila
    list.pop(0)  # elimina fila de datos
    list.pop(0)  # elimina publico en general
    clientes = []
    compras = []
    list.sort(reverse=True)  # ordena lista de mayor a menor
    for j in range(5):  # ciclo para añadir solo 5 mayores a otra lista
        i = list[j]
        clientes.append(i[0])  # añade clientes con mayor monto de compra
        compras.append(i[1])  # añade monto de compras
    plt.figure(figsize=(10, 3))
    plt.barh(compras, clientes)  # grafica con valores de ambas listas
    plt.suptitle('Clientes con mayor aportación ($$$): ')
    plt.show()
    wb.save("base.xlsx")
    wb.close()

def mcusu():  # menu configuración de usuarios
    init(autoreset=True)
    while True:
        print(
            Style.BRIGHT + "Bienvenido a la configuración de usuarios.\n1.-Añadir usuarios\n2.-Eliminar usuarios\n3.-Salir")
        try:
            op = int(input(Style.BRIGHT + Fore.CYAN + "Seleccione una opción: "))
            if op in range(1, 4):
                return op
            else:
                print(Fore.RED + "Error, solo acepta los números del 1 al 3.")
        except ValueError:
            print(Fore.RED + "Error, ingrese solo números")


def anadir_usu():  # funcion para añadir nuevos usuarios para posterior poder ingresar al sistema
    print(Fore.YELLOW + "---Nuevo usuario y contraseña---")
    info = []
    wb.active = 1
    sheet = wb.active
    usu = input("Nuevo usuario: ")
    info.append(usu)
    contra = input("Nueva contraseña: ")
    info.append(contra)
    sheet.append(info)  # añade a excel valores de usuario y contraseña
    print(Fore.YELLOW + "Usuario y contraseña registrados")
    wb.save("base.xlsx")
    wb.close()
    return info


def eliminar_usu():  # eliminar usuuarios existentes
    print(Fore.YELLOW + "---Eliminar usuario---")
    print("Esta es su lista de usuarios: ")
    wb.active = 1
    ws = wb.active  # variable
    print("No.        USUARIOS: ")
    a = 1
    for i in range(2, ws.max_row + 1):  # imprime valores de usuario apartir de la segunda celda
        print(a, end=" ")
        a = a + 1  # contador
        print("     ", ws.cell(row=i, column=1).value)
    while True:
        try:
            print("ESTE NUMERO ES SOLO DE REFERENICA PARA SU BORRADO, NO ES EL DE NO. DE USUARIO REAL")
            borrar = int(input("No. de usuario que deseas borrar: "))
            if borrar in range(1, a):
                ws.delete_rows(borrar + 1)  # borrar usuario seleccionado
                print(Fore.YELLOW + "Borrado exitosamente")
                break
            else:
                print(Fore.RED + "Introduce un número válido.")
        except ValueError:
            print(Fore.RED + "Ingrese solo números.")
    wb.save("base.xlsx")
    wb.close()

# inicio del código
datos = list([])
f = date.today()
init(autoreset=True)
print(Fore.YELLOW + Style.BRIGHT + "***     Sistema de Ventas MAVA     ***")
usu = ""
contra = ""
s = 0
while s != 1:  # depende de lo que regrese la funcion sesion sera si me deja ingresar al sistema
    usu = input(Fore.RED + "USUARIO: ")
    contra = input(Fore.RED + "CONTRASEÑA: ")
    s = sesion(usu, contra)
    if s == 1:
        while True:
            op = menup()
            if op == 1:  # Almacen
                while True:
                    opmca = mca()
                    if opmca == 1:
                        anadir_producto()  # alta de productos
                    elif opmca == 2:
                        eliminar_producto()  # eliminar productos
                    elif opmca == 3:
                        consultar_productos()  # consultar produtcos
                        while True:
                            opap = menu_conprod()
                            if opap == 1:
                                consultar_productos()  # consulta de productos
                            elif opap == 2:
                                print('Gracias')
                                break
                    elif opmca == 4:
                        compra_producto()
                    elif opmca == 5:
                        print('Gracias.')
                        break
            elif op == 2:
                while True:
                    opmdc = mdc()  # Clientes
                    if opmdc == 1:
                        anadir_contacto()  # añadir contactos
                    elif opmdc == 2:
                        eliminar_contacto()  # eliminar
                    elif opmdc == 3:
                        print('Gracias.')
                        break
            elif op == 3:
                while True:
                    opv = mventas()  # Ventas
                    if opv == 1:
                        nota, total = nota_venta()
                        cli = seleccionar_cliente(total)
                        imprimir_nota(nota, cli, total)
                        resta_almacen(nota)
                    elif opv == 2:
                        print('Gracias.')
                        break
            elif op == 4:
                while True:
                    opest = mestadisticas()  # Estadisticas
                    if opest == 1:
                        pvendidod()
                    elif opest == 2:
                        g = ganancias_diaria()
                        guardar_ganancia(g)
                    elif opest == 3:
                        pexistencia()
                    elif opest == 4:
                        cliente_frecuente()
                    elif opest == 5:
                        clientes_dinero()
                    elif opest == 6:
                        print('Gracias.')
                        break
            elif op == 5:
                if usu == "gerente" and contra == "123abc":
                    while True:
                        opmc = mcusu()
                        if opmc == 1:
                            anadir_usu()
                        elif opmc == 2:
                            eliminar_usu()
                        elif opmc == 3:
                            print('Gracias.')
                        break
                else:
                    print(Fore.RED + "No tienes acceso a esa opción")
            elif op == 6:
                print('Gracias')
                print(Fore.YELLOW + Style.BRIGHT + "--- LA BASE HA SIDO ACTUALIZADA ---")
                break
    elif s == 2:
        print("Usuario y/o contraseña incorrectos.")